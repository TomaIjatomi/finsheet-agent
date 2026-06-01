"""
Workbook introspection for the spreadsheet MCP server.

Pure-Python utilities to:
  - List sheets and dimensions in an xlsx
  - Infer schema (headers, columns, dtypes, fund boundaries, sample rows)
  - Load a cell range as a pandas DataFrame (used by the sandbox to prepare
    named ranges for execute_python)
  - Load a cell range as a JSON-friendly structure (used by get_range)

Tolerates the structural complexity present in the FinSheet-Bench-style
files: title rows above the header, multi-line headers, fund-as-row
separators, average summary rows, blank rows between fund blocks.
"""

from __future__ import annotations

import re
from dataclasses import dataclass
from pathlib import Path
from typing import Any

import openpyxl
import pandas as pd
from openpyxl.utils import column_index_from_string, get_column_letter


@dataclass
class ColumnInfo:
    name: str
    col_letter: str
    col_idx: int  # 1-indexed
    dtype: str  # "string" | "number" | "date" | "mixed" | "empty"


@dataclass
class FundBoundary:
    fund: str
    start_row: int  # first data row for this fund (1-indexed)
    end_row: int  # last data row for this fund (inclusive)
    n_companies: int


@dataclass
class SheetSchema:
    name: str
    rows: int
    cols: int
    header_row: int  # 1-indexed row number
    data_start_row: int
    data_end_row: int
    columns: list[ColumnInfo]
    fund_layout: str  # "column" | "row_separator" | "unknown"
    fund_column: str | None  # col letter if fund_layout == "column"
    fund_boundaries: list[FundBoundary]  # populated for both layouts
    average_rows: list[int]  # row indices of summary/average rows
    sample_rows: list[dict[str, Any]]  # first N data rows as dicts

    def to_dict(self) -> dict:
        return {
            "name": self.name,
            "dimensions": {"rows": self.rows, "cols": self.cols},
            "header_row": self.header_row,
            "data_start_row": self.data_start_row,
            "data_end_row": self.data_end_row,
            "columns": [
                {"name": c.name, "col": c.col_letter, "dtype": c.dtype} for c in self.columns
            ],
            "fund_layout": self.fund_layout,
            "fund_column": self.fund_column,
            "fund_boundaries": [
                {
                    "fund": fb.fund,
                    "start_row": fb.start_row,
                    "end_row": fb.end_row,
                    "n_companies": fb.n_companies,
                }
                for fb in self.fund_boundaries
            ],
            "average_rows": self.average_rows,
            "sample_rows": self.sample_rows,
        }


# ---- Sheet listing -------------------------------------------------------


def list_sheets(file_path: Path | str) -> dict[str, dict[str, int]]:
    """List sheets with dimensions. Returns {sheet_name: {rows, cols}}."""
    wb = openpyxl.load_workbook(file_path, data_only=True, read_only=True)
    try:
        out = {}
        for ws in wb.worksheets:
            out[ws.title] = {"rows": ws.max_row or 0, "cols": ws.max_column or 0}
        return out
    finally:
        wb.close()


# ---- Schema inference ----------------------------------------------------


def _find_header_row(ws, max_scan: int = 15) -> int | None:
    """Heuristic: the header row is one that

      (a) contains the cell 'Company' (case-sensitive, exact match after strip), and
      (b) has at least 3 populated cells total.

    Condition (b) rules out title rows like 'Synthetic1 Portfolio Company Data'
    where 'Company' appears inside a single long cell.

    Returns 1-indexed row number, or None if not found.
    """
    max_col = ws.max_column or 0
    for row_idx in range(1, max_scan + 1):
        populated = 0
        company_match = False
        for col_idx in range(1, max_col + 1):
            val = ws.cell(row=row_idx, column=col_idx).value
            if val is None or val == "":
                continue
            populated += 1
            if str(val).strip() == "Company":
                company_match = True
        if company_match and populated >= 3:
            return row_idx
    return None


def _infer_dtype_for_column(ws, col_idx: int, data_start: int, data_end: int) -> str:
    """Inspect ~20 cells in the column to classify its dtype."""
    n_str = n_num = n_date = n_empty = 0
    sampled = 0
    for r in range(data_start, min(data_end + 1, data_start + 20)):
        val = ws.cell(row=r, column=col_idx).value
        sampled += 1
        if val is None or val == "":
            n_empty += 1
            continue
        if isinstance(val, (int, float)):
            n_num += 1
        elif isinstance(val, str):
            # Try ISO date
            if re.match(r"^\d{4}-\d{2}-\d{2}", val.strip()):
                n_date += 1
            else:
                n_str += 1
        else:
            n_str += 1
    if sampled == 0 or n_empty == sampled:
        return "empty"
    non_empty = sampled - n_empty
    if n_num / non_empty > 0.7:
        return "number"
    if n_date / non_empty > 0.7:
        return "date"
    if n_str / non_empty > 0.7:
        return "string"
    return "mixed"


def _detect_fund_layout(
    ws, header_row: int, columns: list[ColumnInfo], data_start: int, data_end: int
) -> tuple[str, str | None]:
    """Return (layout, fund_column_letter_or_None).

    layout is "column" if a 'Fund' column exists, "row_separator" if rows
    appear with only column A populated, "unknown" otherwise.
    """
    fund_col_letter = None
    for c in columns:
        if c.name.strip().lower() == "fund":
            fund_col_letter = c.col_letter
            break
    if fund_col_letter is not None:
        return "column", fund_col_letter

    # Check for row-separator pattern: scan first 30 data rows; if at least
    # one row has column A non-empty and columns B..N all empty, it's a divider.
    for r in range(data_start, min(data_end + 1, data_start + 30)):
        a_val = ws.cell(row=r, column=1).value
        if a_val is None or a_val == "":
            continue
        others_empty = True
        for col_idx in range(2, len(columns) + 1):
            if ws.cell(row=r, column=col_idx).value not in (None, ""):
                others_empty = False
                break
        if others_empty:
            return "row_separator", None
    return "unknown", None


def _scan_fund_boundaries_row_sep(
    ws, columns: list[ColumnInfo], data_start: int, data_end: int
) -> tuple[list[FundBoundary], list[int]]:
    """For row-separator layout: each fund row is a divider; companies follow."""
    boundaries: list[FundBoundary] = []
    average_rows: list[int] = []
    current_fund: str | None = None
    current_start: int | None = None
    current_count = 0

    n_cols = len(columns)
    for r in range(data_start, data_end + 1):
        a_val = ws.cell(row=r, column=1).value
        if a_val is None or a_val == "":
            continue
        a_str = str(a_val).strip()
        # Detect average row
        if a_str.lower().endswith("average") or " average" in a_str.lower():
            average_rows.append(r)
            continue
        # Detect divider row (only col A populated)
        others_empty = all(
            ws.cell(row=r, column=col_idx).value in (None, "") for col_idx in range(2, n_cols + 1)
        )
        if others_empty:
            # Close previous fund
            if current_fund is not None and current_start is not None and current_count > 0:
                boundaries.append(
                    FundBoundary(
                        fund=current_fund,
                        start_row=current_start,
                        end_row=r - 1,
                        n_companies=current_count,
                    )
                )
            current_fund = a_str
            current_start = r + 1
            current_count = 0
        else:
            if current_fund is not None:
                current_count += 1
    # Flush last fund
    if current_fund is not None and current_start is not None and current_count > 0:
        boundaries.append(
            FundBoundary(
                fund=current_fund,
                start_row=current_start,
                end_row=data_end,
                n_companies=current_count,
            )
        )
    return boundaries, average_rows


def _scan_fund_boundaries_column(
    ws, fund_col_letter: str, data_start: int, data_end: int
) -> tuple[list[FundBoundary], list[int]]:
    """For fund-as-column layout: group consecutive rows by Fund column value."""
    fund_col_idx = column_index_from_string(fund_col_letter)
    boundaries: list[FundBoundary] = []
    average_rows: list[int] = []
    current_fund: str | None = None
    current_start: int | None = None
    current_count = 0

    for r in range(data_start, data_end + 1):
        a_val = ws.cell(row=r, column=1).value
        if a_val is not None:
            a_str = str(a_val).strip().lower()
            if a_str.endswith("average") or " average" in a_str:
                average_rows.append(r)
                continue
        fund_val = ws.cell(row=r, column=fund_col_idx).value
        if fund_val is None or fund_val == "":
            continue
        fund_str = str(fund_val).strip()
        if fund_str != current_fund:
            if current_fund is not None and current_start is not None and current_count > 0:
                boundaries.append(
                    FundBoundary(
                        fund=current_fund,
                        start_row=current_start,
                        end_row=r - 1,
                        n_companies=current_count,
                    )
                )
            current_fund = fund_str
            current_start = r
            current_count = 1
        else:
            current_count += 1
    if current_fund is not None and current_start is not None and current_count > 0:
        boundaries.append(
            FundBoundary(
                fund=current_fund,
                start_row=current_start,
                end_row=data_end,
                n_companies=current_count,
            )
        )
    return boundaries, average_rows


def get_sheet_schema(
    file_path: Path | str, sheet: str | None = None, n_sample_rows: int = 3
) -> SheetSchema:
    """Infer schema for a sheet. Sample rows give the LLM a feel for the data."""
    wb = openpyxl.load_workbook(file_path, data_only=True)
    try:
        ws = wb[sheet] if sheet else wb.active
        rows = ws.max_row or 0
        cols = ws.max_column or 0

        header_row = _find_header_row(ws)
        if header_row is None:
            # Fallback to row 1
            header_row = 1
        data_start = header_row + 1
        data_end = rows

        # Build columns from header row
        columns: list[ColumnInfo] = []
        for col_idx in range(1, cols + 1):
            val = ws.cell(row=header_row, column=col_idx).value
            if val is None or str(val).strip() == "":
                continue
            # Multi-line header: strip newlines for the name; keep readable form
            name = re.sub(r"\s+", " ", str(val)).strip()
            dtype = _infer_dtype_for_column(ws, col_idx, data_start, data_end)
            columns.append(
                ColumnInfo(
                    name=name,
                    col_letter=get_column_letter(col_idx),
                    col_idx=col_idx,
                    dtype=dtype,
                )
            )

        fund_layout, fund_col = _detect_fund_layout(ws, header_row, columns, data_start, data_end)
        if fund_layout == "column":
            boundaries, avg_rows = _scan_fund_boundaries_column(ws, fund_col, data_start, data_end)
        elif fund_layout == "row_separator":
            boundaries, avg_rows = _scan_fund_boundaries_row_sep(ws, columns, data_start, data_end)
        else:
            boundaries, avg_rows = [], []

        # Sample rows: first N data rows that aren't dividers or averages
        sample_rows: list[dict[str, Any]] = []
        skip_set = set(avg_rows)
        seen = 0
        for r in range(data_start, data_end + 1):
            if seen >= n_sample_rows:
                break
            if r in skip_set:
                continue
            row_dict: dict[str, Any] = {}
            non_empty = False
            for c in columns:
                val = ws.cell(row=r, column=c.col_idx).value
                if val is not None:
                    non_empty = True
                row_dict[c.name] = val
            # Skip divider rows (mostly empty)
            if non_empty and sum(1 for v in row_dict.values() if v is not None) > 1:
                sample_rows.append(row_dict)
                seen += 1

        return SheetSchema(
            name=ws.title,
            rows=rows,
            cols=cols,
            header_row=header_row,
            data_start_row=data_start,
            data_end_row=data_end,
            columns=columns,
            fund_layout=fund_layout,
            fund_column=fund_col,
            fund_boundaries=boundaries,
            average_rows=avg_rows,
            sample_rows=sample_rows,
        )
    finally:
        wb.close()


# ---- Range loading -------------------------------------------------------

_RANGE_RE = re.compile(r"^([A-Z]+)(\d+):([A-Z]+)(\d+)$")


def _parse_range(range_str: str) -> tuple[int, int, int, int]:
    """Parse 'A4:N156' to (col_start_idx, row_start, col_end_idx, row_end)."""
    m = _RANGE_RE.match(range_str.strip().upper())
    if not m:
        raise ValueError(f"Invalid range: {range_str!r}; expected like 'A4:N156'")
    c1, r1, c2, r2 = m.groups()
    return (
        column_index_from_string(c1),
        int(r1),
        column_index_from_string(c2),
        int(r2),
    )


def get_range_as_dict(file_path: Path | str, sheet: str, range_str: str) -> dict:
    """Read a cell range and return a JSON-friendly structure.

    Returns:
        {
          "range": "A4:D10",
          "rows": [
            {"row": 4, "cells": {"A4": "Company", "B4": "Sector", ...}},
            ...
          ]
        }
    """
    col_s, row_s, col_e, row_e = _parse_range(range_str)
    wb = openpyxl.load_workbook(file_path, data_only=True, read_only=True)
    try:
        ws = wb[sheet]
        out_rows = []
        for r in range(row_s, row_e + 1):
            cells: dict[str, Any] = {}
            for c in range(col_s, col_e + 1):
                cell_ref = f"{get_column_letter(c)}{r}"
                val = ws.cell(row=r, column=c).value
                cells[cell_ref] = val
            out_rows.append({"row": r, "cells": cells})
        return {"range": range_str.upper(), "rows": out_rows}
    finally:
        wb.close()


def load_range_as_df(
    file_path: Path | str, sheet: str, range_str: str, header_in_range: bool = True
) -> pd.DataFrame:
    """Load a cell range as a pandas DataFrame.

    If header_in_range is True, the first row of the range is treated as
    column headers. Otherwise columns are named 'col_1', 'col_2', ...
    """
    col_s, row_s, col_e, row_e = _parse_range(range_str)
    wb = openpyxl.load_workbook(file_path, data_only=True, read_only=True)
    try:
        ws = wb[sheet]
        rows = []
        for r in range(row_s, row_e + 1):
            rows.append([ws.cell(row=r, column=c).value for c in range(col_s, col_e + 1)])
        if header_in_range and rows:
            headers = [str(h) if h is not None else f"col_{i + 1}" for i, h in enumerate(rows[0])]
            # Multi-line header: collapse to single line
            headers = [re.sub(r"\s+", " ", h).strip() for h in headers]
            return pd.DataFrame(rows[1:], columns=headers)
        return pd.DataFrame(rows, columns=[f"col_{i + 1}" for i in range(col_e - col_s + 1)])
    finally:
        wb.close()


# ---- Citation helper -----------------------------------------------------


def format_citation(claim: str, sheet: str, cells: list[str]) -> str:
    """Format a claim with a structured cell-range citation.

    Citations look like: 'Apex has Entry EV $478M [Portfolio!G6]'
    or 'Total unrealized: $4.2B [Portfolio!G6,G14,G27]'
    """
    if not cells:
        return claim
    cells_clean = [c.strip().upper() for c in cells]
    citation = f"[{sheet}!{','.join(cells_clean)}]"
    if claim:
        return f"{claim} {citation}"
    return citation
