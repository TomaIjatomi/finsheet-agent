"""
Spreadsheet serialization for LLM consumption.

Implements the 'default' serialization format used by FinSheet-Bench
(Ravnik et al. 2026, §4.1): pipe-separated rows preserving all cells
including blanks, multi-line header newlines, and fund-divider rows.

This is deliberately a messy text representation — the whole point is
that the LLM has to parse the structural complexity (multi-line headers,
fund dividers, average rows, blank separators) to answer correctly.
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook


def _cell_to_str(value) -> str:
    """Convert an openpyxl cell value to its string representation."""
    if value is None:
        return ""
    if isinstance(value, float):
        # Strip trailing zeros for cleaner output but preserve precision
        if value.is_integer():
            return str(int(value))
        return f"{value:g}"
    return str(value)


def serialize_xlsx(xlsx_path: Path | str, sheet_name: str | None = None) -> str:
    """Serialize an xlsx file to FinSheet-style text format.

    Args:
        xlsx_path: Path to the xlsx file.
        sheet_name: Sheet to serialize. If None, uses the active sheet.

    Returns:
        Text representation with one pipe-separated row per line,
        preserving blanks, multi-line header newlines, fund dividers,
        and other structural complexity.
    """
    wb = load_workbook(xlsx_path, data_only=True, read_only=True)
    ws = wb[sheet_name] if sheet_name else wb.active

    lines: list[str] = []
    for row in ws.iter_rows(values_only=True):
        # Strip trailing all-None tail to avoid spurious blank cells
        # (openpyxl pads to max_column)
        cells = list(row)
        while cells and cells[-1] is None:
            cells.pop()
        if not cells:
            lines.append("")  # preserve blank rows (fund separators)
            continue
        cell_strs = [_cell_to_str(c) for c in cells]
        lines.append("| " + " | ".join(cell_strs) + " |")

    wb.close()
    return "\n".join(lines)


def estimate_tokens(text: str) -> int:
    """Rough token estimate (~4 chars per token for English/numeric text).

    Used for prompt-size sanity checks. Replace with tiktoken or similar
    if you need precision; this is good enough for choosing serialization
    strategy at the file-size level.
    """
    return len(text) // 4
