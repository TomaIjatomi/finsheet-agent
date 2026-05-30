"""
Generates synthetic FinSheet-style spreadsheets from template specs.

Two outputs per template:
  1. A "canonical DataFrame" — the clean, structured underlying data,
     used to compute ground truth deterministically.
  2. An xlsx file — the messy, formatted version the LLM sees, with
     fund dividers, multi-line headers, and other layout complexity.

Following FinSheet-Bench Section 3.1.4: ground truth is computed from
the canonical DataFrame, never from the xlsx file. This keeps the
ground truth deterministic and the spreadsheet free to be as messy as needed.
"""
from __future__ import annotations

import random
from datetime import date, timedelta
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .templates import (
    BOARD_FIRST_NAMES,
    BOARD_LAST_NAMES,
    CITIES,
    COLUMNS,
    COMPANY_PREFIXES,
    COMPANY_SUFFIXES,
    SECTORS,
    TemplateSpec,
)

FONT = "Arial"  # consistent with xlsx skill standard
HEADER_FILL = PatternFill("solid", start_color="DDDDDD")
FUND_DIVIDER_FILL = PatternFill("solid", start_color="F0F0E0")
AVERAGE_ROW_FILL = PatternFill("solid", start_color="EEEEEE")


def _roman(n: int) -> str:
    vals = [(1000,"M"),(900,"CM"),(500,"D"),(400,"CD"),(100,"C"),(90,"XC"),
            (50,"L"),(40,"XL"),(10,"X"),(9,"IX"),(5,"V"),(4,"IV"),(1,"I")]
    out = []
    for v, sym in vals:
        while n >= v:
            out.append(sym)
            n -= v
    return "".join(out)

def _fund_name(idx: int, scheme: str) -> str:
    if scheme == "roman":
        return f"Fund {_roman(idx)}"
    if scheme == "descriptive":
        descriptors = ["Growth", "Income", "Stability", "Diversify", "Equity",
                       "Opportunity", "Yield", "Capital"]
        return descriptors[(idx - 1) % len(descriptors)]
    return f"Fund {chr(64 + idx)}"  # Fund A, B, C, ...

def _company_name(rng: random.Random) -> str:
    return f"{rng.choice(COMPANY_PREFIXES)} {rng.choice(COMPANY_SUFFIXES)}"

def _board_members(rng: random.Random, n: int) -> list[str]:
    return [
        f"{rng.choice(BOARD_FIRST_NAMES)} {rng.choice(BOARD_LAST_NAMES)}"
        for _ in range(n)
    ]


def build_canonical_df(spec: TemplateSpec, seed: int = 42) -> pd.DataFrame:
    """Generate the clean underlying DataFrame for one template.

    Deterministic given (spec, seed). Ground truth derives from this DataFrame.
    """
    rng = random.Random(seed + hash(spec.file_id) % 10000)
    rows = []

    # Allocate companies across funds with realistic uneven distribution
    fund_sizes = []
    remaining = spec.n_companies
    for i in range(spec.n_funds):
        if i == spec.n_funds - 1:
            fund_sizes.append(remaining)
        else:
            avg = remaining / (spec.n_funds - i)
            size = max(3, int(rng.gauss(avg, avg * 0.25)))
            size = min(size, remaining - (spec.n_funds - i - 1) * 3)
            fund_sizes.append(size)
            remaining -= size

    used_names: set[str] = set()
    for fund_idx, fund_size in enumerate(fund_sizes, start=1):
        fund_name = _fund_name(fund_idx, spec.fund_naming)
        # Each successive fund is more recent; vintage drift.
        vintage_year = 2010 + (fund_idx - 1) * 2

        for _ in range(fund_size):
            # Unique company name
            for _attempt in range(100):
                name = _company_name(rng)
                if name not in used_names:
                    used_names.add(name)
                    break

            sector = rng.choice(SECTORS)
            hq = rng.choice(CITIES)
            # Older funds have more realized investments
            realized_prob = max(0.2, 0.85 - (spec.n_funds - fund_idx) * 0.0 + (fund_idx - 1) * 0.0)
            # Simpler: vintage drives realization
            years_since = 2026 - vintage_year
            realized_prob = min(0.9, 0.15 + 0.08 * years_since)
            status = "Realized" if rng.random() < realized_prob else "Unrealized"

            entry_year = vintage_year + rng.randint(0, 3)
            entry_date = date(entry_year, rng.randint(1, 12), rng.randint(1, 28))
            if status == "Realized":
                hold_years = rng.randint(3, 8)
                exit_date = entry_date + timedelta(days=hold_years * 365 + rng.randint(-90, 90))
                if exit_date.year > 2025:
                    exit_date = date(2025, exit_date.month, min(exit_date.day, 28))
            else:
                exit_date = None

            # Financials — entry first, exit derived for realized
            entry_ev = round(rng.uniform(50, 2000), 1)
            ebitda_margin = rng.uniform(0.08, 0.28)
            entry_ebitda = round(entry_ev * ebitda_margin, 1)
            debt_multiple = rng.uniform(0.5, 4.5)
            net_debt_entry = round(entry_ebitda * debt_multiple, 1)
            ownership = round(rng.uniform(0.15, 1.0), 3)

            if status == "Realized":
                multiple = rng.uniform(0.7, 3.5)
                exit_ev = round(entry_ev * multiple, 1)
                exit_ebitda_growth = rng.uniform(0.8, 2.5)
                exit_ebitda = round(entry_ebitda * exit_ebitda_growth, 1)
            else:
                exit_ev = None
                exit_ebitda = None

            board = _board_members(rng, rng.randint(2, 5))

            rows.append({
                "Company": name,
                "Sector": sector,
                "Headquarters": hq,
                "Status": status,
                "Entry Date": entry_date,
                "Exit Date": exit_date,
                "Entry EV": entry_ev,
                "Exit EV": exit_ev,
                "Entry EBITDA": entry_ebitda,
                "Exit EBITDA": exit_ebitda,
                "Net Debt at Entry": net_debt_entry,
                "Ownership %": ownership,
                "Board Members": board,
                "Fund": fund_name,
            })

    return pd.DataFrame(rows)


def _format_header(col: str, multiline: bool) -> str:
    """Apply multi-line splitting for some column names if multiline_headers=True."""
    if not multiline:
        return col
    splits = {
        "Entry Date": "Entry\nDate",
        "Exit Date": "Exit\nDate",
        "Entry EV": "Entry\nEnterprise Value",
        "Exit EV": "Exit\nEnterprise Value",
        "Entry EBITDA": "Entry\nEBITDA",
        "Exit EBITDA": "Exit\nEBITDA",
        "Net Debt at Entry": "Net Debt\nat Entry",
        "Ownership %": "Ownership\n%",
        "Board Members": "Board\nMembers",
        "Headquarters": "HQ",
    }
    return splits.get(col, col)


def _format_cell(value, dtype: str, list_sep: str):
    if value is None:
        return ""
    if dtype == "date":
        return value.isoformat() if hasattr(value, "isoformat") else str(value)
    if dtype == "money":
        return value
    if dtype == "pct":
        return value
    if dtype == "list":
        return list_sep.join(value) if isinstance(value, list) else str(value)
    return str(value) if value is not None else ""


def write_xlsx(spec: TemplateSpec, df: pd.DataFrame, out_path: Path) -> None:
    """Render the canonical DataFrame to an xlsx file following the spec's layout choices."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Portfolio"

    # Top metadata rows
    ws["A1"] = f"{spec.file_id.title()} Portfolio Company Data"
    ws["A1"].font = Font(name=FONT, bold=True, size=14)
    ws["A2"] = "$ in millions"
    ws["A2"].font = Font(name=FONT, italic=True, size=10)
    current_row = 4

    # Determine column list (Fund column dropped from data layout if fund_as_column=False)
    cols_to_render = [(c, t) for c, t in COLUMNS if not (c == "Fund" and not spec.fund_as_column)]

    # Header row
    for col_idx, (col_name, _dtype) in enumerate(cols_to_render, start=1):
        cell = ws.cell(row=current_row, column=col_idx)
        cell.value = _format_header(col_name, spec.multiline_headers)
        cell.font = Font(name=FONT, bold=True)
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(wrap_text=spec.multiline_headers, vertical="center")
    if spec.multiline_headers:
        ws.row_dimensions[current_row].height = 32
    current_row += 1

    # Data rows, fund by fund
    fund_order = df["Fund"].drop_duplicates().tolist()
    for fund_idx, fund in enumerate(fund_order):
        fund_df = df[df["Fund"] == fund].reset_index(drop=True)

        # Fund divider row (when fund is NOT a column)
        if not spec.fund_as_column:
            divider_cell = ws.cell(row=current_row, column=1)
            divider_cell.value = fund
            divider_cell.font = Font(name=FONT, bold=True, size=11)
            divider_cell.fill = FUND_DIVIDER_FILL
            current_row += 1

        for _, row in fund_df.iterrows():
            for col_idx, (col_name, dtype) in enumerate(cols_to_render, start=1):
                ws.cell(row=current_row, column=col_idx).value = _format_cell(
                    row[col_name], dtype, spec.list_separator
                )
                ws.cell(row=current_row, column=col_idx).font = Font(name=FONT)
            current_row += 1

        # Average row (after fund block)
        if spec.has_average_rows:
            ws.cell(row=current_row, column=1).value = f"{fund} Average"
            ws.cell(row=current_row, column=1).font = Font(name=FONT, italic=True, bold=True)
            ws.cell(row=current_row, column=1).fill = AVERAGE_ROW_FILL
            for col_idx, (col_name, dtype) in enumerate(cols_to_render, start=1):
                if dtype == "money" and col_name in (
                    "Entry EV", "Exit EV", "Entry EBITDA", "Exit EBITDA", "Net Debt at Entry"
                ):
                    avg = fund_df[col_name].dropna().mean()
                    if pd.notna(avg):
                        ws.cell(row=current_row, column=col_idx).value = round(float(avg), 1)
                        ws.cell(row=current_row, column=col_idx).font = Font(
                            name=FONT, italic=True
                        )
                        ws.cell(row=current_row, column=col_idx).fill = AVERAGE_ROW_FILL
            current_row += 1

        if spec.blank_rows_between_funds and fund_idx < len(fund_order) - 1:
            current_row += 1  # blank separator row

    # Set column widths
    for col_idx, (col_name, _dtype) in enumerate(cols_to_render, start=1):
        col_letter = get_column_letter(col_idx)
        if col_name == "Board Members":
            ws.column_dimensions[col_letter].width = 36
        elif col_name in ("Company", "Headquarters"):
            ws.column_dimensions[col_letter].width = 22
        else:
            ws.column_dimensions[col_letter].width = 15

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)


def generate_one(spec: TemplateSpec, version: str, out_dir: Path, seed: int = 42) -> Path:
    """Generate one (template, version) xlsx. Returns the output path.

    The canonical DataFrame is also returned via the side-effect of writing
    it as a parquet alongside the xlsx, for downstream ground-truth computation.
    """
    df = build_canonical_df(spec, seed=seed)
    if version != "A":
        from .variants import apply_variant
        df = apply_variant(df, version)
    out_xlsx = out_dir / f"{spec.file_id}_{version}.xlsx"
    # Effective spec adjusts based on version (some variants change structure flags)
    effective_spec = spec
    if version in ("B", "C"):
        from .variants import variant_overrides
        effective_spec = variant_overrides(spec, version)
    write_xlsx(effective_spec, df, out_xlsx)
    # Save canonical df for ground truth use
    df_path = out_dir / f"{spec.file_id}_{version}.canonical.parquet"
    df.to_parquet(df_path, index=False)
    return out_xlsx
